# -*- coding: utf-8 -*-
# @Author  : XQQ
# @Software: PyCharm Community Edition
import configparser,os
import unittest
from openpyxl import load_workbook

from ddt import ddt, data, unpack

from Interface_Auto.conf.read_path import Excel_path, Log_path, Config_path  # 引入路径变量
from Interface_Auto.public.Class_Excel_RW import excel_RW
from Interface_Auto.public.Http_Request import Http_Request
from Interface_Auto.public.class_logging import MyLogger
from Interface_Auto.public.class_mysql import Mysql_Read

cf=configparser.ConfigParser()
cf.read(Config_path,encoding="utf-8")
mode=cf.get("BIDLOAN_FLAG","mode")
case_list=eval(cf.get("BIDLOAN_FLAG","case_list"))

excel_data=excel_RW()
bidloan_data=excel_data.excel_Read(Excel_path,'bidLoan',mode,case_list)
mylogger=MyLogger("mylogger",Log_path)

@ddt
class Bidloan_Request_test(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        print('开始测试')
        cf=configparser.ConfigParser()
        cf.read(Config_path,encoding='utf-8')
        cls.api=cf["HTTP"]["api"]

    @data(*bidloan_data)
    @unpack
    def test_Bidloan_Request(self,id,module,description,interface_address,mobilephone,password,
                              loanId,amount,method,code_name,expect_code,actually_code,db_check,result,reason):
        '''投资接口'''
        workbook = load_workbook(Excel_path)
        # 初始化手机号
        if mobilephone == "$mobilephone":
            mobilephone = workbook["mobile"].cell(row=2, column=1).value-1
            workbook.save(Excel_path)
            memberId=Mysql_Read().mysql_read("select m.Id from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
        else:
            memberId=mobilephone
        url = self.api + interface_address
        params={"memberId":memberId,"password":password,"loanId":loanId,"amount":amount}
        if method=='get':
            try:
                t=Http_Request().get_Request(url,params)
                mylogger.mylog().info('用例编号为'+str(id)+'投资接口GET请求测试,用户ID:'+str(memberId))
                excel_data.excel_Write(Excel_path,'bidLoan',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code) #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'reason', t.json()["msg"])
                #投资成功后，校验数据库数据
                if expect_code=='10001':
                    result=Mysql_Read().mysql_read("select m.Amount from invest m where m.MemberID='" + str(memberId) + "'")[0]
                    if int(result) ==int(amount): #投资成功，查询用户ID，应该有投资记录
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试数据库结果错误，用例失败')
                # 投资失败后，校验数据库数据
                else:
                    result = Mysql_Read().mysql_read("select m.Amount from invest m where m.MemberID='" + str(memberId) + "'")
                    if result == None: #投资失败，查询用户ID，应该没有投资记录
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试数据库结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('GET请求测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id)+1,'result', 'fail')
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'reason', t.json()["msg"])
                raise e
        elif method=='post':
            try:
                t=Http_Request().post_Request(url,params)
                mylogger.mylog().info('用例标号'+str(id)+'投资接口POST请求测试,手机号:'+str(mobilephone))
                excel_data.excel_Write(Excel_path,'bidLoan',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code) #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'reason', t.json()["msg"])
                # 投资成功后，校验数据库数据
                if expect_code == '10001':
                    result = Mysql_Read().mysql_read(
                        "select m.Amount from invest m where m.MemberID='" + str(memberId) + "'")[0]
                    if int(result) == int(amount):  # 投资成功，查询用户ID，应该有投资记录
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试数据库结果错误，用例失败')
                # 投资失败后，校验数据库数据
                else:
                    result = Mysql_Read().mysql_read("select m.Amount from invest m where m.MemberID='" + str(memberId) + "'")
                    if result == None: #投资失败，查询用户ID，应该没有投资记录
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试数据库结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('POST测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'result', 'fail')
                excel_data.excel_Write(Excel_path, 'bidLoan', int(id) + 1, 'reason', t.json()["msg"])
                raise e

    @classmethod
    def tearDownClass(cls):
        print('测试结束')