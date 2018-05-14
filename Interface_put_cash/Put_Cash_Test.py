# -*- coding: utf-8 -*-
# @Time    : 2018/4/13 10:18
# @Author  : XQQ
# @Software: PyCharm Community Edition
import configparser
import unittest

from ddt import ddt, data, unpack
from openpyxl import load_workbook
from Interface_Auto.conf.read_path import Excel_path, Log_path, Config_path  # 引入路径变量
from Interface_Auto.public.Class_Excel_RW import excel_RW
from Interface_Auto.public.Http_Request import Http_Request
from Interface_Auto.public.class_logging import MyLogger
from Interface_Auto.public.class_mysql import Mysql_Read

cf=configparser.ConfigParser()
cf.read(Config_path,encoding="utf-8")
mode=cf.get("WITHDRAW_FLAG","mode")
case_list=eval(cf.get("WITHDRAW_FLAG","case_list"))

excel_data=excel_RW()
put_cash_Data=excel_data.excel_Read(Excel_path,'withdraw',mode,case_list)
mylogger=MyLogger("mylogger",Log_path)
@ddt
class Put_Cash_test(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        print('开始测试')
        cf=configparser.ConfigParser()
        cf.read(Config_path,encoding='utf-8')
        cls.api=cf["HTTP"]["api"]

    @data(*put_cash_Data)
    @unpack
    def test_Put_Cash(self,id,module,description,interface_address,mobilephone,amount,
                              method,code_name,expect_code,actually_code,db_check,result,reason):
        '''提现接口'''
        workbook=load_workbook(Excel_path)
        if mobilephone=="$mobilephone":
            mobilephone=int(workbook["mobile"].cell(row=2,column=1).value)-1
        workbook.save(Excel_path)
        url=self.api+interface_address
        params={"mobilephone":mobilephone,"amount":amount}
        if method=='get':
            try:
                LeaveAmount_old = Mysql_Read().mysql_read(
                    "select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
                t=Http_Request().get_Request(url,params)
                mylogger.mylog().info('用例编号为'+str(id)+'提现接口GET请求测试,手机号：'+str(mobilephone)+',当前余额为：'+str(LeaveAmount_old))
                excel_data.excel_Write(Excel_path,'withdraw',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code) #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'reason', t.json()["msg"])
                # 提现成功后，校验数据库数据
                if expect_code == '10001':
                    # 校验提现前后的余额
                    LeaveAmount_new=Mysql_Read().mysql_read(
                        "select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
                    mylogger.mylog().info('用例编号为'+str(id)+'提现接口GET请求测试,手机号：' + str(mobilephone) + ',提现后余额为：' + str(LeaveAmount_new))
                    if float(LeaveAmount_new)==float(LeaveAmount_old)-float(amount):
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试结果错误，用例失败')
                # 提现失败后，校验数据库数据
                else:
                    #校验提现前后的余额
                    LeaveAmount_new = Mysql_Read().mysql_read(
                        "select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
                    mylogger.mylog().info('用例编号为'+str(id)+'提现接口GET请求测试,手机号：' + str(mobilephone) + ',充值后余额为：' + str(LeaveAmount_new))
                    if float(LeaveAmount_new) == float(LeaveAmount_old):
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('GET请求测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'withdraw', int(id)+1,'result', 'fail')
                excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'reason', t.json()["msg"])
                raise e
        elif method=='post':
            try:
                LeaveAmount_old=Mysql_Read().mysql_read("select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")
                t=Http_Request().post_Request(url,params)
                mylogger.mylog().info('用例编号为'+str(id)+'充值接口POST请求测试,手机号：' + str(mobilephone) + ',当前余额为：' + str(LeaveAmount_old))
                excel_data.excel_Write(Excel_path,'withdraw',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code)  #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'reason', t.json()["msg"])
                # 提现成功后，校验数据库数据
                if expect_code == '10001':
                    # 校验提现前后的余额
                    LeaveAmount_new = Mysql_Read().mysql_read(
                        "select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
                    mylogger.mylog().info('用例编号为'+str(id)+'充值接口POST请求测试,手机号：' + str(mobilephone) + ',充值后余额为：' + str(LeaveAmount_new))
                    if float(LeaveAmount_new) == float(LeaveAmount_old) - float(amount):
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试结果错误，用例失败')
                # 提现失败后，校验数据库数据
                else:
                    # 校验提现前后的余额
                    LeaveAmount_new = Mysql_Read().mysql_read(
                        "select m.LeaveAmount from member m where m.MobilePhone='" + str(mobilephone) + "'")[0]
                    mylogger.mylog().info('用例编号为'+str(id)+'充值接口POST请求测试,手机号：' + str(mobilephone) + ',充值后余额为：' + str(LeaveAmount_new))
                    if float(LeaveAmount_new) == float(LeaveAmount_old):
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('POST测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'result', 'fail')
                excel_data.excel_Write(Excel_path, 'withdraw', int(id) + 1, 'reason', t.json()["msg"])
                raise e

    @classmethod
    def tearDownClass(cls):
        print('测试结束')