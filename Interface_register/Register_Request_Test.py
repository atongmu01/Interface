# -*- coding: utf-8 -*-
# @Time    : 2018/4/13 10:18
# @Author  : XQQ
# @Software: PyCharm Community Edition
import configparser,os
import unittest
from openpyxl import load_workbook

from ddt import ddt, data, unpack

from Interface_Auto.conf.read_path import Excel_path, Log_path, Config_path  # 引入路径变量
from Interface_Auto.public.Class_Excel_RW import excel_RW
from Interface_Auto.public.Http_Request import Http_Request
from Interface_Auto.public.class_logging import MyLogger
from Interface_Auto.public.class_mysql import Mysql_Read

cf=configparser.ConfigParser()
cf.read(Config_path,encoding="utf-8")
mode=cf.get("REGISTER_FLAG","mode")
case_list=eval(cf.get("REGISTER_FLAG","case_list"))

excel_data=excel_RW()
register_Data=excel_data.excel_Read(Excel_path,'register',mode,case_list)
mylogger=MyLogger("mylogger",Log_path)

@ddt
class Register_Request_test(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        print('开始测试')
        cf=configparser.ConfigParser()
        cf.read(Config_path,encoding='utf-8')
        cls.api=cf["HTTP"]["api"]

    @data(*register_Data)
    @unpack
    def test_Register_Request(self,id,module,description,interface_address,mobilephone,pwd,
                              regname,method,code_name,expect_code,actually_code,db_check,result,reason):
        '''注册接口'''
        workbook=load_workbook(Excel_path)
        #初始化手机号
        if mobilephone=="$mobilephone":
            mobilephone=workbook["mobile"].cell(row=2,column=1).value
            workbook.save(Excel_path)
        url=self.api+interface_address
        params={"mobilephone":mobilephone,"pwd":pwd,"regname":regname}
        if method=='get':
            try:
                t=Http_Request().get_Request(url,params)
                mylogger.mylog().info('用例编号为'+str(id)+'注册接口GET请求测试,手机号:'+str(mobilephone))
                excel_data.excel_Write(Excel_path,'register',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code) #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'reason', t.json()["msg"])
                #注册成功后，更新初始化手机号，校验数据库数据
                if expect_code=='10001':
                    excel_data.excel_Write(Excel_path, 'mobile', 2, '$mobilephone', int(mobilephone)+1)
                    excel_data.excel_Write(Excel_path, 'mobile', 2, '$pwd', pwd)
                    result=Mysql_Read().mysql_read("select * from member m where m.MobilePhone='" + str(mobilephone) + "'")
                    if result != None: #注册成功，查询手机号应该有数据
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试数据库结果错误，用例失败')
                # 注册失败后，校验数据库数据
                else:
                    result = Mysql_Read().mysql_read("select * from member m where m.MobilePhone='" + str(mobilephone) + "'")
                    if result == None: #注册失败，查询手机号应该为None
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('GET请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('GET请求测试数据库结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('GET请求测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'register', int(id)+1,'result', 'fail')
                excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'reason', t.json()["msg"])
                raise e
        elif method=='post':
            try:
                t=Http_Request().post_Request(url,params)
                mylogger.mylog().info('用例标号'+str(id)+'注册接口POST请求测试,手机号:'+str(mobilephone))
                excel_data.excel_Write(Excel_path,'register',int(id)+1,'actually_code',t.json()[code_name])
                self.assertEqual(t.json()[code_name],expect_code) #先检查接口返回数据是否通过
                excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'reason', t.json()["msg"])
                # 注册成功后，更新初始化手机号，校验数据库数据
                if expect_code == '10001':
                    excel_data.excel_Write(Excel_path, 'mobile', 2, '$mobilephone', int(mobilephone) + 1)
                    excel_data.excel_Write(Excel_path, 'mobile', 2, '$pwd', pwd)
                    result = Mysql_Read().mysql_read(
                        "select * from member m where m.MobilePhone='" + str(mobilephone) + "'")
                    if result != None:  # 注册成功，查询手机号应该有数据
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试数据库结果错误，用例失败')
                # 注册失败后，校验数据库数据
                else:
                    result = Mysql_Read().mysql_read(
                        "select * from member m where m.MobilePhone='" + str(mobilephone) + "'")
                    if result == None:  # 注册失败，查询手机号应该为None
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'pass')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'pass')
                        mylogger.mylog().info('POST请求测试数据库结果正确，用例通过')
                    else:
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'db_check', 'fail')
                        excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'fail')
                        mylogger.mylog().info('POST请求测试数据库结果错误，用例失败')
            except Exception as e:
                mylogger.mylog().error('POST测试结果错误，用例失败')
                excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'result', 'fail')
                excel_data.excel_Write(Excel_path, 'register', int(id) + 1, 'reason', t.json()["msg"])
                raise e

    @classmethod
    def tearDownClass(cls):
        print('测试结束')

if __name__ == '__main__':
    suite = unittest.TestSuite()
    suite.addTest(unittest.TestLoader().loadTestsFromTestCase(Register_Request_test))
    runner=unittest.TextTestRunner()
    runner.run(suite)