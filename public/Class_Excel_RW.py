#coding=utf-8
#Author: Xqq
import os
from openpyxl import Workbook
from openpyxl import load_workbook

class excel_RW:
    def excel_Create(self,excel_path,sheet_name):
        workbook=Workbook()
        workbook.create_sheet(sheet_name)
        workbook.save(excel_path)

    def excel_Write(self,excel_path,sheet_name,row_no,col_name,data):
        if not os.path.exists(excel_path):
            self.excel_Create(excel_path,sheet_name)
        workbook=load_workbook(excel_path)
        sheet=workbook[sheet_name]
        for k in range(sheet.max_column):  #获取写入第几列
            if sheet.cell(row= 1, column=k + 1).value==col_name:
                col_no=k+1
                break
        sheet.cell(row=row_no,column=col_no).value=data
        workbook.save(excel_path)

    def excel_Read(self,excel_path,sheet_name,mode,case_list):
        if not os.path.exists(excel_path):
            self.excel_Create(excel_path,sheet_name)
        workbook = load_workbook(excel_path)
        sheet = workbook[sheet_name]
        result=[]
        for i in range(1,sheet.max_row):
            dict_1={}
            for k in range(sheet.max_column):
                dict_1[(sheet.cell(row=1,column=k+1).value)]=(sheet.cell(row=i+1,column=k+1).value)
            result.append(dict_1)

        interface_Data = []
        if mode == '1':
            for item in result:
                if int(item["id"]) in case_list:
                    interface_Data.append(item)
        if mode == '0':
            interface_Data = result
        workbook.save(excel_path)
        return interface_Data


if __name__=='__main__':
    data=[  ['学号','姓名','性别','班级'],
            ['0501','喵酱','女','python5'],
            ['0506','老宝儿','女','python5'],
            ['0511','杰东','男','python5'],
            ['0518','呼呼','男','python5'],
            ['0519','时光','男','python5'],
            ['0521','米拉','女','python5'],
            ['0522','西红柿','男','python5'],
            ['0523','小小土豆','男','python5'],
            ['0524','索尔','男','python5'],
            ['0525','车厘子','男','python5'],
            ['0526','心心','女','python5'],
            ['0527','小白','男','python5'],
            ['0529','木木','男','python5'],
            ['0530','a.I','男','python5'],
            ['0532','朱元璋','男','python5'],
            ['0539','佐佐','女','python5'],
            ['0541','Eva','女','python5'],
            ['0548','飘瑶','女','python5'],
            ['0554','樱子','女','python5'],
            ['0559','海芋','男','python5'] ]
    excel_1=excel_RW()
    excel_1.excel_Write('python5.xlsx','python5',data)
    result=excel_1.excel_Read('python5.xlsx','python5')
    print(result)